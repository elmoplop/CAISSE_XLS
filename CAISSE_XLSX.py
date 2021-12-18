#!/usr/bin/env python3

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import formatting, styles
import datetime
import os, re

#Getting
files = []
for file in os.listdir("."):
    if file.endswith(".html"):
        files.append(file)

if len(files) > 1:
    print('More than one html files has been found:')
    for file in files:
        print(file)
    htmlReport = input('\nWhich file do you want to process ?\n')
else:
    htmlReport = files[0]

#Interpretation de l'html par Beautifull soup
soup = BeautifulSoup(open(htmlReport), "html.parser", from_encoding='ISO-8859-15')

### Recuperation du titre et stockage de la date de rapport
date = str(soup.find("h1").get_text().split("- ")[1].replace(" ","_"))

### Recuperation des tableau interressant dans le html ###
# Récuperation de la table resume info générales
generalTable = soup.find("table", {"class":"display mainStdDisp"})

# Récuperation des tables article par rayons
articles = soup.find_all(string=re.compile("Articles par rayon"))
articlesTablesList = []
for article in articles:
    articlesTablesList.append(article.find_next("table"))

articleTable = []

#### Suppression des header de tableau et marge en tableau unique
i = 0
for table in articlesTablesList:
    j = 0
    for line in table:
        if j != 0:
            articleTable.append(line)
        if i == 0 and j == 0:
            articleTable.append(line)
        j = j + 1
    i = i + 1

# Récuperation de la table TVA
TVATable = soup.find("h2").find_next("table")

# Récuperation de la table Rayons
storeTable = soup.find("h2", string="Rayons").find_next("table")

### Recuperation des tableau interressant dans le html
# i=0
# for table in soup.find_all("table"):
#     i = i+1
   # print("\n{}".format(i))

    # if i == 1:  # Table d'information générale
    #     generalTable = table
#     if i == 2:  # Table récapitulatif TVA
#         TVATable = table
#     if i == 5:  # Table recapitulatif tout rayon
#         storeTable = table
#     if i == 7:
#         articleTable = table  # Table détaillé des ventes par tout rayon
#
### Transfomation en liste pour les articles
allSales = []
for line in articleTable:  # Pour chacun ligne de la table
    saleList = []
    for elem in line:  # Pour chacun des éléments de la ligne
        saleList.append(elem.text.encode("windows-1252"))  # On ne garde que le texte brut que l'on ré-encode et on ajoute a la liste decrivant la vente
    allSales.append(saleList)  # On construit la liste de liste des ventes
# print(allSales)

# Modification de la liste atricles
salesHeader = allSales[0] # Récupération du header
del allSales[0]  # Suppression du header dans la liste
i = 2
for line in allSales:
    line[4] = float(line[4].decode("utf-8").replace('€','').replace(',','.'))  # passage en float d'un prix
    line[5] = float(line[5].decode("utf-8").replace('€','').replace(',','.'))  # passage en float d'un prix
    line[6] = float(line[6].decode("utf-8").replace('€','').replace(',','.'))  # passage en float d'un prix
    line[7] = float(line[7].decode("utf-8").replace('€','').replace(',','.'))  # passage en float d'un prix
    line[8] = float(line[8].decode("utf-8").replace('€','').replace(',','.'))  # passage en float d'un prix
    i = i + 1

# Transfomation en liste pour les info generale
allGeneral = []
for line in generalTable:
    generalList = []
    for elem in line:
        generalList.append(elem.text.encode("windows-1252"))
    allGeneral.append(generalList)

# Transfomation en liste pour la TVA
allTVA = []
for line in TVATable:
    TVAList = []
    for elem in line:
        TVAList.append(elem.text.encode("windows-1252"))
    allTVA.append(TVAList)

# Transfomation en liste pour les rayons
allStore = []
for line in storeTable:
    storeList = []
    for elem in line:
        storeList.append(elem.text.encode("windows-1252"))
    allStore.append(storeList)

# Modification de la liste rayons
storeHeader = allStore[0]  # Récupération du header
storeHeader = storeHeader + ['%Commision', 'A facturer', 'Date Mail', ' Date Facture', 'Date Virement']  # Ajout de collonnes
del allStore[0]  # Suppression du header dans la liste
i = 2
for line in allStore:
    line[2] = float(line[2].decode("utf-8").replace('€','').replace(',','.')) #passage en float d'un prix
    line[3] = float(line[3].decode("utf-8").replace('€','').replace(',','.')) #passage en float d'un prix
    line[4] = float(line[4].decode("utf-8").replace('€','').replace(',','.')) #passage en float d'un prix
    line.append(30)  # Ajout du pourcentage de commission
    formula = '=E' + str(i) + '-(E' + str(i) + '*F' + str(i) + '/100)'  # Ajout de la formule calculant le total à facturer
    line.append(formula)
    i = i + 1

# Calcul des totaux et ajout a la liste
allTots = []
formula = '=SUM(E2:E' + str(i-1) + ')'
allTots.append(['', 'Total Chiffre:', formula, ])
formula = '=SUM(G2:G' + str(i-1) + ')'
allTots.append(['', 'Total Facturé:', formula, ])
formula = '=C' + str(i+1) + '-C'+ str(i+2)
allTots.append(['', 'Total bénéfice asso:', formula, ])

#Creation du XLSX
now = datetime.datetime.now()
genpath = 'GEN_' + now.strftime("%Y-%m-%d_%H%M%S") + '/'

os.mkdir(genpath)
wb = Workbook()
dest_filename = genpath + 'ALL_' + date + '.xlsx'

#Feuillet rayon
ws1 = wb.active
ws1.title = "Général"

#Mise en forme
ws1.column_dimensions['A'].width = 35
titleStyle = Font(size="15", bold=True)

ws1.append([("Rapport Mensuel les 3 créatrices et Cie " + " - " + date.replace("_","/") ), ])
ws1.cell(row=1, column=1,).font = titleStyle
ws1.append(["", ])

for elem in allGeneral:
    ws1.append(elem)

ws1.append([" ", ])
ws1.append(["TVA", ])


for elem in allTVA:
    ws1.append(elem)

#Feuillet rayon
ws2 = wb.create_sheet(title="Récapitulatif Rayon")
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 20

ws2.append(storeHeader) #Injection du header

for cell in ws2["1:1"]: #Passage de la premier ligne en gras
    cell.font = Font(bold=True)

i = 2
for elem in allStore:
    ws2.append(elem)
    i = i + 1

ws2.append([" ", ])

for tot in allTots: #Injection des totaux
    ws2.append(tot)

#Changement de format pour les totaux
ws2['{}{}'.format("C", i + 1)].number_format = '0.00€'
ws2['{}{}'.format("C", i + 2)].number_format = '0.00€'
ws2['{}{}'.format("C", i + 3)].number_format = '0.00€'


#Changement de format pour les colonne de ws2
for row in range(2, ws2.max_row+1):
    ws2['{}{}'.format("D", row)].number_format = '0.00€'
for row in range(2, ws2.max_row+1):
    ws2['{}{}'.format("E", row)].number_format = '0.00€'
for row in range(2, ws2.max_row+1):
    ws2['{}{}'.format("G", row)].number_format = '0.00€'

#Feuillets Articles par rayon
def close_store():
    """Mthode permettant d'ajouter le total a la fin d'un feuillet individuel et de rensigner la comparaison avec le feuillet rayon"""
    if store:
        j = 2
        for line in allStore:
            # print('{}_{}'.format(line[0], elem[1]))
            if line[0] == store or (line[0].decode() == "Sans" and store.decode() == "NA"):
                storeTot = line[4]
                cellTr = 'E' + str(j)

            j = j + 1

        formula = '=SUM(I2:I' + str(i - 1) + ')'
        saleTots = ['', '', '', '', '', '', '', 'Total Chiffre:', formula, ]
        red_color = 'ffc7ce'
        red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
        cellRef = '\'' + store.decode() + '\'!I' + str(i + 1)
        ws2.conditional_formatting.add(cellTr,
                                       formatting.rule.CellIsRule(operator='notEqual', formula=[cellRef], fill=red_fill))
        wsx.append([" ", ])
        wsx.append(saleTots)
        # Changement de format pour les colonne de wsx
        for row in range(2, ws2.max_row + 1):
            wsx['{}{}'.format("F", row)].number_format = '0.00€'
        for row in range(2, ws2.max_row + 1):
            wsx['{}{}'.format("H", row)].number_format = '0.00€'
        for row in range(2, ws2.max_row + 1):
            wsx['{}{}'.format("I", row)].number_format = '0.00€'

store = ""
i = 2
for elem in allSales:
    if elem[1] != store: #Si on change de rayon dans la liste
        close_store()
        i = 2
        store = elem[1]
        # print(store)
        wsx = wb.create_sheet(title=elem[1].decode())
        wsx.column_dimensions['A'].width = 30
        wsx.column_dimensions['B'].width = 30
        wsx.column_dimensions['C'].width = 15
        wsx.append(salesHeader) #Injection du header
        for cell in wsx["1:1"]:  # Passage de la premiere ligne en gras
            cell.font = Font(bold=True)
    wsx.append(elem)
    i = i + 1
close_store()

wb.save(filename=dest_filename)

print('{} has been created'.format(dest_filename))

######################################################


#Feuillets Articles par rayon
def close_file():
    """Methode permettant d'ajouter le total à un fichier puis de l'enregistrer"""
    formula = '=SUM(I2:I' + str(i - 1) + ')'
    saleTots = ['', '', '', '', '', '', '', 'Total Chiffre:', formula, 'Commission:', 30, '%', ]
    # com = ['', '', '', '', '', '', '', 'Commission:', 30, '%', ]
    formula = '=I' + str(i + 1) + '-(I' + str(i + 1) + '*K' + str(i + 1) + '/100)'
    fact = ['', '', '', '', '', '', '', 'A facturer:', formula, ]
    wsy.append([" ", ])
    wsy.append(saleTots)
    # wsy.append(com)
    wsy.append(fact)
    # Changement de format pour les colonne de wsx
    for row in range(2, ws2.max_row + 1):
        wsy['{}{}'.format("F", row)].number_format = '0.00€'
    for row in range(2, ws2.max_row + 1):
        wsy['{}{}'.format("H", row)].number_format = '0.00€'
    for row in range(2, ws2.max_row + 1):
        wsy['{}{}'.format("I", row)].number_format = '0.00€'
    wb.save(filename=dest_filename)
    print('{} has been created'.format(dest_filename))


store = ""
i = 2
for elem in allSales:
    # print(elem[1], store)
    if elem[1].decode() != store: #Si on change de rayon dans la liste
        if store:
            close_file()
        i = 2
        store = elem[1].decode()
        wb = Workbook()
        dest_filename = genpath + str(store.replace(' ','_')) + '_' + date + '.xlsx'
        wsy = wb.active
        wsy.title = store
        wsy.column_dimensions['A'].width = 30
        wsy.column_dimensions['B'].width = 30
        wsy.column_dimensions['C'].width = 15
        wsy.append(salesHeader) #Injection du header
        for cell in wsx["1:1"]:  # Passage de la premiere ligne en gras
            cell.font = Font(bold=True)
    wsy.append(elem)
    i = i + 1
close_file()


input('Press Enter to exit')
#print(allSales)